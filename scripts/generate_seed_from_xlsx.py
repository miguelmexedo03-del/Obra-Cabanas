"""
Converte Cabanas_Checklist.xlsx num ficheiro SQL de seed para a Supabase.

USO:
    python3 scripts/generate_seed_from_xlsx.py \
        --checklist ../Cabanas_Checklist.xlsx \
        --out supabase/migrations/0002_seed_checklist.sql

Output:
- INSERTs para a tabela `divisoes` (uma por apartamento × divisão única)
- INSERTs para a tabela `elementos` (um por linha do checklist classificável)
- INSERTs para a tabela `tarefas_gantt` (144 rows: 24 pais + 120 fases)
"""
from openpyxl import load_workbook
from pathlib import Path
import argparse


FASE_MAP = {
    1: 'Tetos',
    2: 'Paredes',
    3: 'Carpintaria',
    4: 'Chão e Rodapé',
    5: 'WC Equipamentos',
}


def classify_fase(elemento):
    if not elemento:
        return None
    e = str(elemento).lower()
    if "teto" in e:
        return 1
    if "parede" in e:
        return 2
    if any(k in e for k in ["aro", "porta", "móvei", "movei", "banca", "eletrodomésti", "eletrodomesti", "bancada"]):
        return 3
    if "chão" in e or "chao" in e or "rodapé" in e or "rodape" in e:
        return 4
    if any(k in e for k in ["lavatório", "lavatorio", "sanita", "chuveiro", "duche", "toalheiro"]):
        return 5
    return None


def sql_escape(s):
    """Escape single quotes for SQL string literals."""
    if s is None:
        return 'null'
    s = str(s).replace("'", "''")
    return f"'{s}'"


def parse_checklist(path):
    wb = load_workbook(path, data_only=True)
    rows = []
    divisoes_set = {}  # (apartamento_id, nome) -> ordem
    for ap in range(1, 25):
        name = f"AP{ap}"
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        current_divisao = None
        current_elemento = None
        divisao_counter = 0
        for r in range(3, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e_notes = ws.cell(row=r, column=5).value
            f_resp = ws.cell(row=r, column=6).value

            if a is not None and str(a).strip():
                current_divisao = str(a).strip()
                key = (ap, current_divisao)
                if key not in divisoes_set:
                    divisao_counter += 1
                    divisoes_set[key] = divisao_counter
            if b is not None and str(b).strip():
                current_elemento = str(b).strip()

            has_b = b is not None and str(b).strip()
            has_c = c is not None and str(c).strip()
            if not (has_b or has_c):
                continue
            fase = classify_fase(current_elemento)
            if fase is None:
                continue

            rows.append({
                "ap": ap,
                "divisao": current_divisao or "—",
                "elemento": current_elemento or "",
                "sub_elemento": str(c).strip() if has_c else None,
                "fase": fase,
                "checked": 1 if (d is not None and "✓" in str(d)) else 0,
                "notas": str(e_notes).strip() if e_notes else None,
                "responsavel": str(f_resp).strip() if f_resp else None,
            })
    return rows, divisoes_set


def generate_sql(rows, divisoes_set, out_path):
    lines = []
    lines.append("-- Seed gerado automaticamente por generate_seed_from_xlsx.py")
    lines.append("-- Não editar à mão. Re-gerar se o checklist original mudar.")
    lines.append("")
    lines.append("begin;")
    lines.append("")

    # --- Divisões ---
    lines.append("-- Divisões (por apartamento)")
    lines.append("insert into divisoes (apartamento_id, nome, ordem) values")
    divisao_values = []
    # Manter ordem determinística
    for (ap, nome), ordem in sorted(divisoes_set.items()):
        divisao_values.append(f"  ({ap}, {sql_escape(nome)}, {ordem})")
    lines.append(",\n".join(divisao_values) + ";")
    lines.append("")

    # --- Elementos ---
    lines.append("-- Elementos (items individuais do checklist)")
    lines.append("-- Nota: concluído=false por defeito, o seed histórico não traz estado real.")
    lines.append("-- Se quiseres importar o estado atual, ver script sync_checklist.py.")
    lines.append("")
    # Chunk em batches de 500 para ficar legível
    BATCH = 500
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i+BATCH]
        lines.append("insert into elementos (apartamento_id, divisao_id, fase_id, elemento, sub_elemento, concluido, notas, responsavel) values")
        values = []
        for row in batch:
            # divisao_id resolvido via subquery — mais robusto a rearranjos
            divisao_sql = f"(select id from divisoes where apartamento_id = {row['ap']} and nome = {sql_escape(row['divisao'])} limit 1)"
            values.append(
                f"  ({row['ap']}, {divisao_sql}, {row['fase']}, "
                f"{sql_escape(row['elemento'])}, {sql_escape(row['sub_elemento'])}, "
                f"{'true' if row['checked'] else 'false'}, "
                f"{sql_escape(row['notas'])}, {sql_escape(row['responsavel'])})"
            )
        lines.append(",\n".join(values) + ";")
        lines.append("")

    # --- Tarefas Gantt: 24 pais + 120 filhos ---
    lines.append("-- Tarefas Gantt: 24 pais (apartamento) + 120 filhos (fases)")
    lines.append("-- Datas ficam null; serão preenchidas depois via LoB ou UI.")
    lines.append("")
    lines.append("-- Pais (nivel=1)")
    parent_values = []
    for ap in range(1, 25):
        parent_values.append(
            f"  (null, {ap}, null, 1, 'AP{ap} — Obra Cabanas', 'por_fazer')"
        )
    lines.append("insert into tarefas_gantt (parent_id, apartamento_id, fase_id, nivel, nome, status) values")
    lines.append(",\n".join(parent_values) + ";")
    lines.append("")

    lines.append("-- Filhos (nivel=2, 5 fases por apartamento)")
    lines.append("-- Usa subquery para obter parent_id do apartamento correspondente")
    for ap in range(1, 25):
        for fase_id in range(1, 6):
            fase_nome = FASE_MAP[fase_id]
            lines.append(
                f"insert into tarefas_gantt (parent_id, apartamento_id, fase_id, nivel, nome, status) values "
                f"((select id from tarefas_gantt where apartamento_id = {ap} and nivel = 1), "
                f"{ap}, {fase_id}, 2, {sql_escape(fase_nome)}, 'por_fazer');"
            )
    lines.append("")

    lines.append("commit;")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"OK: {out_path} ({len(rows)} elementos, {len(divisoes_set)} divisões, 144 tarefas gantt)")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--checklist", required=True, help="Path para Cabanas_Checklist.xlsx")
    p.add_argument("--out", required=True, help="Output SQL file")
    args = p.parse_args()

    rows, divisoes = parse_checklist(Path(args.checklist))
    generate_sql(rows, divisoes, Path(args.out))
